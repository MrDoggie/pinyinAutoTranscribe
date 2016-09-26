from openpyxl import load_workbook
from openpyxl import Workbook
import re


###
#enter the excel filename here 
###
wb = load_workbook (filename = 'data.xlsx')

ws = wb.active

rows = ws.rows # get row numbers 


###
#enter the input and outputs' col numbers. Only need to change the col number here, no other  adjustment needs to be made.
#make sure the col number entered is correct or it might replace the data you have 
#A->1 B->2 and so on
###
firstInput = 8 #for example, the first pinyin input is in col 8
firstOnset = 9 
firstGlide = 10
firstRhyme = 11
#col 12 is for tone
firstDMGlide = 13
firstDMRhyme = 14

secondInput = 19
secondOnset = 20
secondGlide = 21
secondRhyme = 22
#col 23 is for tone
secondDMGlide = 24
secondDMRhyme =25


#fuyin options: bpmfdtnlkghjqxzcsr; have rule out impossible combination

def transcribe (traditionInput, inputCol, onsetCol, glideCol, rhymeCol, DMGlideCol, DMRhymeCol):
	m_i1 = re.search('^([zhcsrz]+)i$' , traditionInput) # -i after zh, ch, sh,r,z,c,s
	m_i2 = re.search('^([bpmdtnljqxy])i$',traditionInput) #-i after 0, b, p,m,d,t,n,j,q,x; y will be handled later 
	m_a = re.search('^([bpmfdtnlkghzcs]*)a$',traditionInput) #-a, need to consider a along
	m_o = re.search('^([bpmfl]*)o$', traditionInput) #-o, need to consider o along
	m_e = re.search('^([mdtnlkghzcsr]*)e$',traditionInput) #-e, need to consider e along
	m_ai = re.search('^([bpmdtnlkghzcs]*)ai$', traditionInput) #-ai, need to consider ai along
	m_ei = re.search('^([bpmfdtnlkghsz]*)ei$',traditionInput) #-ei, need to consider ei along
	m_ao = re.search('^([bpmdtnlkghzcsr]*)ao$',traditionInput) #-ao, need to consider ao along
	m_ou = re.search('^([pmfdtlkghzcs]*)ou$',traditionInput) #-ou, need to consider ou along
	m_an = re.search('^([bpmfdtnlkghzcsr]*)an$',traditionInput) #-an, need to consider an along
	m_en = re.search('^([bpmfdnkghzcsr]*)en$',traditionInput)#-en, need to consider en along 
	m_ang = re.search('^([bpmfdtnlkghzcsr]*)ang$', traditionInput) #-ang
	m_eng = re.search('^([bpmfdtnlkghzcsr]*)eng$',traditionInput) #-eng
	m_er = re.search('^er$', traditionInput) #er, no other possibility
	m_ia = re.search('^([dljqx])ia$', traditionInput) #-ia ** TO DO y special case
	m_ie = re.search('^([bpmdtnljqx])ie$', traditionInput) #-ie 
	m_iu = re.search('^([mdnljqx])iu$',traditionInput) #-iu
	m_ian = re.search('^([bpmdtnljqx])ian$',traditionInput) #-ian
	m_in = re.search('^([bpmnljqxy])in$',traditionInput) #-in; y will be handled in the swtich branch
	m_iang = re.search('^([bnljqx])iang$', traditionInput) #-iang
	m_ing = re.search('^([bpmdtnljqxy])ing$', traditionInput) #-ing; y will be handled in the switch branch
	m_u = re.search('^([bpmfdtnlkghzcsrw]+)u$', traditionInput) #-ing (ju qu xu yu are exception); wu will be handled in the swtich branch
	m_ua = re.search('^([kghsz]+)ua$',traditionInput) #-ua
	m_uo = re.search('^([dtnlkghzcsr]+)uo$',traditionInput) #-uo
	m_uai = re.search('^([kghzcs]+)uai$',traditionInput) #-uai
	m_ui = re.search('^([dtkghzcsr]+)ui$',traditionInput) #-ui
	m_uan = re.search('^([dtnlkghzcsr]+)uan$',traditionInput) #-uan
	m_un = re.search('^([dtnlkghzcsr]+)un$',traditionInput) #-un
	m_uang = re.search('^([kghzcs]+)uang$',traditionInput) #-uang
	m_ong = re.search('^([dtnlkghzcsr]+)ong$',traditionInput) #-ong
	m_U = re.search('^([jqxy])[uv]$',traditionInput) #yu sound after jqx; yu handled in the switch branch
	m_Ue = re.search('^([jqxy])[uv]e$',traditionInput) #yu sound in ue; yue handled in the switch branch
	m_Uan = re.search('^([jqxy])[uv]an$',traditionInput) #yu sound in uan; yuan handled in the switch branch
	m_Un = re.search('^([jqxy])[uv]n$',traditionInput) #yu sound in un; yun handled in the switch branch
	m_iong = re.search('^([jqx])iong$',traditionInput) #-iong
	m_iao = re.search('^([bpmdtnljqx])iao$', traditionInput) #-iao
	m_v = re.search('^([ln]v)$', traditionInput) # lv and nv
	#special cases
	m_ya = re.search('^ya$', traditionInput) #ya special case
	m_ye = re.search('^ye$', traditionInput) #ye special case
	m_you = re.search('^you$', traditionInput) #you special case
	m_yan = re.search('^yan$', traditionInput) #yan SC
	m_yang = re.search('^yang$', traditionInput) #yang SC
	m_wa = re.search('^wa$', traditionInput) #wa SC
	m_wo = re.search('^wo$', traditionInput) #wo SC
	m_wai = re.search('^wai$', traditionInput) #wai SC
	m_wei = re.search('^wei$', traditionInput) #wei SC
	m_wan = re.search('^wan$', traditionInput) #wan SC
	m_wen = re.search('^wen$', traditionInput) #wen SC
	m_wang = re.search('^wang$', traditionInput) #wang SC
	m_yong = re.search('^yong$',traditionInput) #yong SC
	m_yao = re.search('^yao$', traditionInput) #yao SC
	#switch branches
	if m_i1:
		ws.cell(row = rn+1, column = onsetCol).value=m_i1.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='r'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='i'
	elif m_i2:
		ws.cell(row = rn+1, column = onsetCol).value=m_i2.group(1)
		if m_i2.group(1) == 'y':
			ws.cell(row = rn+1, column = onsetCol).value = '0' #if input is yi, output 0 0 i 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='i'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='i'
	elif m_a:
		ws.cell(row = rn+1, column = onsetCol).value=m_a.group(1) 
		if m_a.group(1) == '': #consider a along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='a'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='a'
	elif m_o:
		ws.cell(row = rn+1, column = onsetCol).value=m_o.group(1) 
		if m_o.group(1) == '': #consider o along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='o'
		ws.cell(row = rn+1, column = DMGlideCol).value='w'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_e:
		ws.cell(row = rn+1, column = onsetCol).value=m_e.group(1) 
		if m_e.group(1) == '': #consider e along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='e'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_ai:
		ws.cell(row = rn+1, column = onsetCol).value=m_ai.group(1) 
		if m_ai.group(1) == '': #consider ai along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ai'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ai'
	elif m_ei:
		ws.cell(row = rn+1, column = onsetCol).value=m_ei.group(1) 
		if m_ei.group(1) == '': #consider ei along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ei'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ei'
	elif m_ao:
		ws.cell(row = rn+1, column = onsetCol).value=m_ao.group(1) 
		if m_ao.group(1) == '': # ao along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ao'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='au'
	elif m_ou:
		ws.cell(row = rn+1, column = onsetCol).value=m_ou.group(1) 
		if m_ou.group(1) == '': # ou along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ou'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='eu'
	elif m_an:
		ws.cell(row = rn+1, column = onsetCol).value=m_an.group(1) 
		if m_an.group(1) == '': # an along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_en:
		ws.cell(row = rn+1, column = onsetCol).value=m_en.group(1) 
		if m_en.group(1) == '': #en along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='en'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='en'
	elif m_wen:  #wen SC
		ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='en'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='en'
	elif m_ang:
		ws.cell(row = rn+1, column = onsetCol).value=m_ang.group(1) 
		if m_ang.group(1) == '': #ang along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ang'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ang'
	elif m_eng:
		ws.cell(row = rn+1, column = onsetCol).value=m_eng.group(1) 
		if m_eng.group(1) == '': #eng along
			ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='eng'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='eng'
	elif m_er: #only possible pinyin is er. Therefore no onset
		ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='er'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='er'
	elif m_ia:
		ws.cell(row = rn+1, column = onsetCol).value=m_ia.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='a'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='a'
	elif m_ya: #deal with ya special case
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='a'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='a'
	elif m_ie: 
		ws.cell(row = rn+1, column = onsetCol).value=m_ie.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='e'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_ye: #ye special case
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='e'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_iu:
		ws.cell(row = rn+1, column = onsetCol).value=m_iu.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ou'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='eu'
	elif m_you: #you special case
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ou'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='eu'
	elif m_ian:
		ws.cell(row = rn+1, column = onsetCol).value=m_ian.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_yan:
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_in:
		ws.cell(row = rn+1, column = onsetCol).value=m_in.group(1)
		if m_in.group(1) =='y': #handle yin special case
			ws.cell(row=rn+1, column = onsetCol).value = '0' 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='in'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='in'
	elif m_iang:
		ws.cell(row = rn+1, column = onsetCol).value=m_iang.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ang'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ang'
	elif m_yang:
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ang'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ang'
	elif m_ing:
		ws.cell(row = rn+1, column = onsetCol).value=m_ing.group(1) 
		if m_ing.group(1) == 'y': #ying  SC
			ws.cell(row =rn+1, column=onsetCol).value = '0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ing'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='eng'
	elif m_u:
		ws.cell(row = rn+1, column = onsetCol).value=m_u.group(1)
		if m_u.group(1) =='w': #wu SC
			ws.cell(row=rn+1, column=onsetCol).value ='0' 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='u'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='u'
	elif m_ua:
		ws.cell(row = rn+1, column = onsetCol).value=m_ua.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='a'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='a'
	elif m_wa: #handle wa SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='a'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='a'
	elif m_uo:
		ws.cell(row = rn+1, column = onsetCol).value=m_uo.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='o'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_wo: #handle wo SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='o'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_uai:
		ws.cell(row = rn+1, column = onsetCol).value=m_uai.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ai'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ai'
	elif m_wai: #handle wai SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ai'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ai'
	elif m_ui:
		ws.cell(row = rn+1, column = onsetCol).value=m_ui.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ei'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ei'
	elif m_wei: #handle wei SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ei'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ei'
	elif m_uan:
		ws.cell(row = rn+1, column = onsetCol).value=m_uan.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_wan: #handle wan SC
		ws.cell(row = rn+1, column = onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_un:
		ws.cell(row = rn+1, column = onsetCol).value=m_un.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='un'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='en'
	elif m_uang:
		ws.cell(row = rn+1, column = onsetCol).value=m_uang.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ang'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ang'
	elif m_wang: #handle wang SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='u'
		ws.cell(row = rn+1, column = rhymeCol).value='ang'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ang'
	elif m_ong:
		ws.cell(row = rn+1, column = onsetCol).value=m_ong.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='ong'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ong'
	elif m_U:
		ws.cell(row = rn+1, column = onsetCol).value=m_U.group(1)
		if m_U.group(1) == 'y': #yu SC
			ws.cell(row = rn+1, column =onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='v'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='v'
	elif m_Ue:
		ws.cell(row = rn+1, column = onsetCol).value=m_Ue.group(1)
		if m_Ue.group(1)=='y': #yue SC
			ws.cell(row = rn+1, column =onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='v'
		ws.cell(row = rn+1, column = rhymeCol).value='e'
		ws.cell(row = rn+1, column = DMGlideCol).value='v'
		ws.cell(row = rn+1, column = DMRhymeCol).value='e'
	elif m_Uan:
		ws.cell(row = rn+1, column = onsetCol).value=m_Uan.group(1)
		if m_Uan.group(1) =='y': #yuan SC
			ws.cell(row=rn+1, column=onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='v'
		ws.cell(row = rn+1, column = rhymeCol).value='an'
		ws.cell(row = rn+1, column = DMGlideCol).value='v'
		ws.cell(row = rn+1, column = DMRhymeCol).value='an'
	elif m_Un:
		ws.cell(row = rn+1, column = onsetCol).value=m_Un.group(1) 
		if m_Un.group(1)=='y': #yun SC
			ws.cell(row=rn+1, column =onsetCol).value='0'
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='vn'
		ws.cell(row = rn+1, column = DMGlideCol).value='u'
		ws.cell(row = rn+1, column = DMRhymeCol).value='in'
	elif m_iong:
		ws.cell(row = rn+1, column = onsetCol).value=m_iong.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ong'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ong'
	elif m_yong: #yong SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ong'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='ong'
	elif m_iao:
		ws.cell(row = rn+1, column = onsetCol).value=m_iao.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ao'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='au'
	elif m_yao: #yao SC
		ws.cell(row = rn+1, column = onsetCol).value='0' 
		ws.cell(row = rn+1, column = glideCol).value='i'
		ws.cell(row = rn+1, column = rhymeCol).value='ao'
		ws.cell(row = rn+1, column = DMGlideCol).value='i'
		ws.cell(row = rn+1, column = DMRhymeCol).value='au'
	elif m_v:
		ws.cell(row = rn+1, column = onsetCol).value=m_v.group(1) 
		ws.cell(row = rn+1, column = glideCol).value='0'
		ws.cell(row = rn+1, column = rhymeCol).value='v'
		ws.cell(row = rn+1, column = DMGlideCol).value='0'
		ws.cell(row = rn+1, column = DMRhymeCol).value='v'
	else:
		#if you see output ###, it means either the pinyin input is incorrect or our code has bug.
		ws.cell(row = rn+1, column = onsetCol).value='###' 
		ws.cell(row = rn+1, column = glideCol).value='###'
		ws.cell(row = rn+1, column = rhymeCol).value='###'
		ws.cell(row = rn+1, column = DMGlideCol).value='###'
		ws.cell(row = rn+1, column = DMRhymeCol).value='###'
	wb.save('data.xlsx')

for rn in range(len(rows)):
	if rn != 0: #do not transcribe the first row
		traditionInput=str(ws.cell(row=rn+1, column = firstInput).value)
		transcribe(traditionInput, firstInput, firstOnset, firstGlide, firstRhyme, firstDMGlide, firstDMRhyme)

for rn in range(len(rows)):
	if rn != 0: #do not transcribe the first row 
		traditionInput = str(ws.cell(row=rn+1, column = secondInput).value)
		transcribe(traditionInput, secondInput, secondOnset, secondGlide, secondRhyme, secondDMGlide, secondDMRhyme)
